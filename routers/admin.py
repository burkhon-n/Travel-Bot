from fastapi import APIRouter, Request, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from pathlib import Path
from sqlalchemy.orm import Session
from sqlalchemy import func
from io import BytesIO
from datetime import datetime

from database import get_db
from config import Config
from models.Trip import Trip, TripStatus
from models.TripMember import TripMember, PaymentStatus
from webapp_security import require_telegram_webapp

import logging


router = APIRouter(prefix="/admin", tags=["admin"])
BASE_DIR = Path(__file__).resolve().parent.parent
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))


def _require_admin(request: Request) -> int:
    """Lightweight admin check using client-provided header with Telegram ID.
    Returns the tg_id if allowed; raises HTTPException otherwise.
    The WebApp front-end should set X-Telegram-Id from tg.initDataUnsafe.user.id.
    """
    try:
        tg_id = int(request.headers.get("X-Telegram-Id", "0"))
    except Exception:
        tg_id = 0
    if tg_id not in set(Config.ADMINS or []):
        raise HTTPException(status_code=403, detail="Admin access required")
    return tg_id


@router.get("")
async def admin_home(request: Request):
    """Admin dashboard listing trips with quick stats.
    Only accessible from Telegram WebApp for security.
    All mutating API routes are admin-protected via headers.
    """
    # Check if request is from Telegram WebApp
    error_response = require_telegram_webapp(request, Config.BOT_USERNAME)
    if error_response:
        return error_response
    
    db_gen = get_db()
    db: Session = next(db_gen)
    try:
        trips = db.query(Trip).order_by(Trip.id.desc()).all()
        logging.info("admin.dashboard render trips=%s", len(trips))
        trip_rows = []
        for t in trips:
            total = db.query(func.count(TripMember.id)).filter(TripMember.trip_id == t.id).scalar() or 0
            half = db.query(func.count(TripMember.id)).filter(
                TripMember.trip_id == t.id, 
                TripMember.payment_status == PaymentStatus.half_paid
            ).scalar() or 0
            full = db.query(func.count(TripMember.id)).filter(
                TripMember.trip_id == t.id, 
                TripMember.payment_status == PaymentStatus.full_paid
            ).scalar() or 0
            # Half-paid and full-paid both reserve seats
            paid = half + full
            seats_available = None if t.participant_limit is None else max(t.participant_limit - paid, 0)
            trip_rows.append({
                "id": t.id,
                "name": t.name,
                "group_id": t.group_id,
                "status": t.status.value if hasattr(t.status, 'value') else str(t.status),
                "participant_limit": t.participant_limit,
                "price": t.price if t.price is not None else 0,
                "registered": total,
                "half_paid": half,
                "full_paid": full,
                "paid": paid,
                "seats": seats_available,
            })

        return templates.TemplateResponse(
            "admin_dashboard.html",
            {
                "request": request,
                "admins": Config.ADMINS or [],
                "trips": trip_rows,
            }
        )
    finally:
        try:
            next(db_gen)
        except StopIteration:
            pass


@router.get("/trip/{trip_id}")
async def admin_trip_detail(request: Request, trip_id: int):
    """Trip detail page: members and actions.
    Only accessible from Telegram WebApp for security.
    API calls are also protected with admin checks.
    """
    # Check if request is from Telegram WebApp
    error_response = require_telegram_webapp(request, Config.BOT_USERNAME)
    if error_response:
        return error_response
    
    db_gen = get_db()
    db: Session = next(db_gen)
    try:
        trip = db.query(Trip).filter(Trip.id == trip_id).first()
        if not trip:
            raise HTTPException(status_code=404, detail="Trip not found")

        members = db.query(TripMember).filter(TripMember.trip_id == trip.id).order_by(TripMember.joined_at.desc()).all()
        member_rows = []
        
        # Calculate statistics
        total_registered = len(members)
        half_paid_count = 0
        full_paid_count = 0
        not_paid_count = 0
        
        for m in members:
            # Count payment statuses
            if m.payment_status == PaymentStatus.half_paid:
                half_paid_count += 1
            elif m.payment_status == PaymentStatus.full_paid:
                full_paid_count += 1
            elif m.payment_status == PaymentStatus.not_paid:
                not_paid_count += 1
            
            # Fetch user information
            from models.User import User
            user = db.query(User).filter(User.id == m.user_id).first()
            
            # Build full name
            full_name = "Unknown User"
            telegram_id = None
            if user:
                name_parts = []
                if user.first_name:
                    name_parts.append(user.first_name)
                if user.last_name:
                    name_parts.append(user.last_name)
                full_name = " ".join(name_parts) if name_parts else user.email.split('@')[0] if user.email else f"User {m.user_id}"
                telegram_id = user.telegram_id
            
            member_rows.append({
                "id": m.id,
                "user_id": m.user_id,
                "full_name": full_name,
                "telegram_id": telegram_id,
                "payment_status": m.payment_status.value if hasattr(m.payment_status, 'value') else str(m.payment_status),
                "receipt": m.payment_receipt_file_id,
                "joined_at": m.joined_at,
            })
        
        # Calculate total paid (half + full)
        total_paid = half_paid_count + full_paid_count
        
        # Calculate available seats
        seats_available = None
        if trip.participant_limit is not None:
            seats_available = max(trip.participant_limit - total_paid, 0)

        logging.info("admin.trip detail render trip_id=%s members=%s", trip.id, len(member_rows))
        return templates.TemplateResponse(
            "admin_trip.html",
            {
                "request": request,
                "admins": Config.ADMINS or [],
                "trip": trip,
                "members": member_rows,
                "stats": {
                    "registered": total_registered,
                    "half_paid": half_paid_count,
                    "full_paid": full_paid_count,
                    "not_paid": not_paid_count,
                    "total_paid": total_paid,
                    "seats_available": seats_available,
                }
            }
        )
    finally:
        try:
            next(db_gen)
        except StopIteration:
            pass


@router.post("/api/trip/{trip_id}/status")
async def update_trip_status(request: Request, trip_id: int):
    _require_admin(request)
    data = await request.json()
    new_status = (data.get("status") or "").lower()
    if new_status not in ("active", "completed", "cancelled"):
        raise HTTPException(status_code=400, detail="Invalid status")

    db_gen = get_db()
    db: Session = next(db_gen)
    try:
        trip = db.query(Trip).filter(Trip.id == trip_id).first()
        if not trip:
            raise HTTPException(status_code=404, detail="Trip not found")
        # Map string to enum (TripStatus has: active, completed, cancelled)
        trip.status = TripStatus[new_status]
        db.commit()
        logging.info("admin.trip status updated trip_id=%s status=%s", trip_id, new_status)
        return JSONResponse({"ok": True})
    finally:
        try:
            next(db_gen)
        except StopIteration:
            pass


@router.post("/api/invite-links/{trip_id}/regenerate")
async def regenerate_invite_links(request: Request, trip_id: int):
    _require_admin(request)
    from bot import bot as tg_bot

    db_gen = get_db()
    db: Session = next(db_gen)
    try:
        trip = db.query(Trip).filter(Trip.id == trip_id).first()
        if not trip:
            raise HTTPException(status_code=404, detail="Trip not found")
        if not trip.group_id:
            raise HTTPException(status_code=400, detail="Trip has no group_id")

        # Create a direct join link for paid participants
        try:
            p_link = await tg_bot.create_chat_invite_link(
                chat_id=trip.group_id,
                name=f"Participants - {trip.name}",
                creates_join_request=False,
            )
            trip.participant_invite_link = p_link.invite_link if hasattr(p_link, 'invite_link') else getattr(p_link, 'invite_link', None)
        except Exception as e:
            logging.error(f"Failed to create participant link: {e}")
            raise HTTPException(status_code=502, detail="Failed to create participant link")

        # Create a join-request link for guests
        try:
            g_link = await tg_bot.create_chat_invite_link(
                chat_id=trip.group_id,
                name=f"Guests - {trip.name}",
                creates_join_request=True,
            )
            trip.guest_invite_link = g_link.invite_link if hasattr(g_link, 'invite_link') else getattr(g_link, 'invite_link', None)
        except Exception as e:
            logging.error(f"Failed to create guest link: {e}")
            raise HTTPException(status_code=502, detail="Failed to create guest link")

        db.commit()
        logging.info("admin.trip links regenerated trip_id=%s group_id=%s", trip_id, trip.group_id)
        return JSONResponse({
            "ok": True,
            "participant_invite_link": trip.participant_invite_link,
            "guest_invite_link": trip.guest_invite_link,
        })
    finally:
        try:
            next(db_gen)
        except StopIteration:
            pass


@router.post("/api/member/{member_id}/status")
async def update_member_status(request: Request, member_id: int):
    _require_admin(request)
    from bot import bot as tg_bot
    from models.User import User
    
    data = await request.json()
    new_status = (data.get("status") or "").lower()
    if new_status not in ("not_paid", "half_paid", "full_paid"):
        raise HTTPException(status_code=400, detail="Invalid status")

    db_gen = get_db()
    db: Session = next(db_gen)
    try:
        member = db.query(TripMember).filter(TripMember.id == member_id).first()
        if not member:
            raise HTTPException(status_code=404, detail="Member not found")
        
        # Get user and trip info for notification
        user = db.query(User).filter(User.id == member.user_id).first()
        trip = db.query(Trip).filter(Trip.id == member.trip_id).first()
        
        # Store old status for comparison
        old_status = member.payment_status
        
        # Update status
        new_status_obj = PaymentStatus[new_status]
        member.payment_status = new_status_obj
        db.commit()
        
        # Handle auto-kick when status changes to "Not Paid"
        if new_status_obj == PaymentStatus.not_paid and user and trip and trip.group_id:
            try:
                # Try to kick from group
                await tg_bot.ban_chat_member(trip.group_id, user.telegram_id)
                await tg_bot.unban_chat_member(trip.group_id, user.telegram_id, only_if_banned=True)
                logging.info("admin.member auto-kicked member_id=%s tg_id=%s group_id=%s", 
                           member_id, user.telegram_id, trip.group_id)
            except Exception as e:
                logging.error(f"Failed to kick user {user.telegram_id} from group {trip.group_id}: {e}")
                # Continue even if kick fails (user might not be in group)
        
        # Send notification to user
        if user and trip:
            try:
                status_emoji = {
                    PaymentStatus.not_paid: '❌',
                    PaymentStatus.half_paid: '🟡',
                    PaymentStatus.full_paid: '✅'
                }
                status_text = {
                    PaymentStatus.not_paid: 'Not Paid',
                    PaymentStatus.half_paid: 'Half Paid (50%)',
                    PaymentStatus.full_paid: 'Full Paid (100%)'
                }
                
                emoji = status_emoji.get(new_status_obj, '•')
                text = status_text.get(new_status_obj, new_status)
                
                # Build notification message
                msg = f"{emoji} <b>Payment Status Updated</b>\n\n"
                msg += f"🎫 <b>Trip:</b> {trip.name}\n"
                msg += f"💳 <b>New Status:</b> {text}\n\n"
                
                if new_status_obj == PaymentStatus.not_paid:
                    msg += (
                        "<b>⚠️ Your payment status has been reset to Not Paid.</b>\n\n"
                        "<b>Important:</b>\n"
                        "• Your seat is NO LONGER reserved\n"
                        "• You have been removed from the trip group\n"
                        "• Please make payment as soon as possible\n"
                        "• Send your receipt to secure your spot again\n\n"
                    )
                    if trip.card_info:
                        msg += f"💳 <b>Payment Info:</b>\n{trip.card_info}\n\n"
                    half_price = trip.price // 2
                    msg += f"💵 <b>Minimum Payment (50%):</b> {half_price:,} UZS\n"
                    msg += f"💰 <b>Full Price:</b> {trip.price:,} UZS"
                    
                elif new_status_obj == PaymentStatus.half_paid:
                    msg += (
                        "<b>🎉 Your seat is now RESERVED!</b>\n\n"
                        "<b>What's Next:</b>\n"
                        "• You're confirmed for the trip!\n"
                        "• Complete the remaining 50% before departure\n"
                    )
                    
                    # Send group link if status changed from not_paid
                    if old_status == PaymentStatus.not_paid and trip.participant_invite_link:
                        msg += f"\n🔗 <b>Join the Trip Group:</b>\n<a href='{trip.participant_invite_link}'>👉 Click here to join</a>\n\n"
                    elif trip.participant_invite_link:
                        msg += f"• Trip group: <a href='{trip.participant_invite_link}'>Join Here</a>\n\n"
                    else:
                        msg += "• Trip group link coming soon!\n\n"
                    
                    remaining = trip.price // 2
                    msg += f"💵 <b>Remaining Payment:</b> {remaining:,} UZS\n"
                    msg += f"💰 <b>Total Trip Price:</b> {trip.price:,} UZS"
                    
                elif new_status_obj == PaymentStatus.full_paid:
                    msg += (
                        "<b>✅ You're FULLY PAID!</b>\n\n"
                        "<b>🎉 Congratulations!</b>\n"
                        "• All payments completed\n"
                        "• You're all set for the trip\n"
                    )
                    
                    # Send group link if status changed from not_paid
                    if old_status == PaymentStatus.not_paid and trip.participant_invite_link:
                        msg += f"\n🔗 <b>Join the Trip Group:</b>\n<a href='{trip.participant_invite_link}'>👉 Click here to join</a>\n\n"
                    elif trip.participant_invite_link:
                        msg += f"• Trip group: <a href='{trip.participant_invite_link}'>Join Here</a>\n\n"
                    else:
                        msg += "• Stay tuned for trip updates!\n\n"
                    
                    msg += f"💰 <b>Total Paid:</b> {trip.price:,} UZS\n\n"
                    msg += "See you on the trip! 🌍✈️"
                
                await tg_bot.send_message(
                    user.telegram_id,
                    msg,
                    parse_mode='HTML',
                    disable_web_page_preview=True
                )
                logging.info("admin.member status notification sent member_id=%s status=%s tg_id=%s", 
                           member_id, new_status, user.telegram_id)
                
            except Exception as e:
                logging.error(f"Failed to send status update notification to user {user.telegram_id}: {e}")
                # Don't fail the request if notification fails
        
        # Email sending temporarily disabled - will be re-enabled after admin discussion
        # TODO: Add email confirmation when ready
        
        logging.info("admin.member status updated member_id=%s old_status=%s new_status=%s", 
                    member_id, old_status.value if old_status else None, new_status)
        return JSONResponse({"ok": True})
    finally:
        try:
            next(db_gen)
        except StopIteration:
            pass


@router.post("/api/member/{member_id}/kick")
async def kick_member(request: Request, member_id: int):
    _require_admin(request)
    from bot import bot as tg_bot
    from models.User import User

    db_gen = get_db()
    db: Session = next(db_gen)
    try:
        member = db.query(TripMember).filter(TripMember.id == member_id).first()
        if not member:
            raise HTTPException(status_code=404, detail="Member not found")
        trip = db.query(Trip).filter(Trip.id == member.trip_id).first()
        if not trip:
            raise HTTPException(status_code=404, detail="Trip not found")
        user = db.query(User).filter(User.id == member.user_id).first()
        if not user or not trip.group_id:
            raise HTTPException(status_code=400, detail="Missing group or user")

        try:
            await tg_bot.ban_chat_member(trip.group_id, user.telegram_id)
            await tg_bot.unban_chat_member(trip.group_id, user.telegram_id, only_if_banned=True)
        except Exception as e:
            logging.error(f"Failed to kick user {user.telegram_id}: {e}")
            raise HTTPException(status_code=502, detail="Failed to kick user")

        logging.info("admin.member kicked member_id=%s group_id=%s", member_id, trip.group_id)
        return JSONResponse({"ok": True})
    finally:
        try:
            next(db_gen)
        except StopIteration:
            pass


@router.get("/api/trip/{trip_id}/export-excel")
async def export_trip_excel(request: Request, trip_id: int, send_to_chat: bool = False):
    """Export trip members to a styled Excel file and optionally send to Telegram chat."""
    _require_admin(request)
    from models.User import User
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    from bot import bot
    
    db_gen = get_db()
    db: Session = next(db_gen)
    
    try:
        # Get trip and members
        trip = db.query(Trip).filter(Trip.id == trip_id).first()
        if not trip:
            raise HTTPException(status_code=404, detail="Trip not found")
        
        members = (
            db.query(TripMember, User)
            .join(User, User.id == TripMember.user_id)
            .filter(TripMember.trip_id == trip_id)
            .order_by(TripMember.joined_at)
            .all()
        )
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Trip Members"
        
        # Define styles
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Payment status colors
        status_colors = {
            "not_paid": "FEE2E2",      # Red
            "half_paid": "FEF3C7",     # Yellow
            "full_paid": "D1FAE5"      # Green
        }
        
        border_style = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )
        
        # Title row
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"🎫 {trip.name} - Member List"
        title_cell.font = Font(bold=True, size=16, color="1F2937")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        ws.row_dimensions[1].height = 30
        
        # Info row
        ws.merge_cells('A2:H2')
        info_cell = ws['A2']
        info_cell.value = f"💰 Price: {trip.price:,} UZS  |  📅 Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        info_cell.font = Font(size=10, color="6B7280")
        info_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20
        
        # Headers
        headers = ["#", "Full Name", "Telegram ID", "Email", "Payment Status", "Joined Date", "Receipt", "Amount Paid"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style
        
        ws.row_dimensions[4].height = 25
        
        # Data rows
        for idx, (member, user) in enumerate(members, 1):
            row = idx + 4
            
            # Row data
            full_name = f"{user.first_name or ''} {user.last_name or ''}".strip()
            status_text = member.payment_status.value.replace('_', ' ').title()
            joined_date = member.joined_at.strftime('%Y-%m-%d %H:%M') if member.joined_at else ""
            has_receipt = "✅ Yes" if member.payment_receipt_file_id else "❌ No"
            
            # Calculate amount paid based on status
            if member.payment_status == PaymentStatus.full_paid:
                amount_paid = f"{trip.price:,} UZS"
            elif member.payment_status == PaymentStatus.half_paid:
                amount_paid = f"{trip.price // 2:,} UZS"
            else:
                amount_paid = "0 UZS"
            
            row_data = [
                idx,
                full_name,
                user.telegram_id,
                user.email,
                status_text,
                joined_date,
                has_receipt,
                amount_paid
            ]
            
            # Fill row
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col_num)
                cell.value = value
                cell.border = border_style
                cell.alignment = Alignment(horizontal="left" if col_num in [2, 4] else "center", vertical="center")
                
                # Color based on payment status
                status_color = status_colors.get(member.payment_status.value, "FFFFFF")
                cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
                
                # Make payment status bold
                if col_num == 5:
                    cell.font = Font(bold=True)
            
            ws.row_dimensions[row].height = 20
        
        # Summary row
        summary_row = len(members) + 6
        ws.merge_cells(f'A{summary_row}:D{summary_row}')
        summary_cell = ws[f'A{summary_row}']
        
        total_count = len(members)
        half_paid_count = sum(1 for m, u in members if m.payment_status == PaymentStatus.half_paid)
        full_paid_count = sum(1 for m, u in members if m.payment_status == PaymentStatus.full_paid)
        not_paid_count = total_count - half_paid_count - full_paid_count
        
        summary_cell.value = f"📊 Total: {total_count} | ❌ Not Paid: {not_paid_count} | 🟡 Half Paid: {half_paid_count} | ✅ Full Paid: {full_paid_count}"
        summary_cell.font = Font(bold=True, size=11)
        summary_cell.alignment = Alignment(horizontal="center", vertical="center")
        summary_cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        ws.row_dimensions[summary_row].height = 25
        
        # Adjust column widths
        column_widths = [5, 25, 15, 30, 18, 20, 12, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Generate filename
        filename = f"{trip.name.replace(' ', '_')}_Members_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        # Send to chat if requested
        if send_to_chat:
            try:
                # Get admin's Telegram ID from request header
                tg_id = request.headers.get('X-Telegram-Id')
                if tg_id:
                    tg_id = int(tg_id)
                    # Create a copy of the file for sending
                    excel_file_copy = BytesIO(excel_file.getvalue())
                    excel_file_copy.name = filename
                    
                    # Send document to chat
                    await bot.send_document(
                        tg_id,
                        excel_file_copy,
                        caption=f"📊 <b>{trip.name}</b>\n\nMember list export\n📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                        parse_mode='HTML'
                    )
                    logging.info("admin.export_excel sent_to_chat trip_id=%s tg_id=%s", trip_id, tg_id)
            except Exception as e:
                logging.error(f"Failed to send Excel to chat: {e}")
        
        logging.info("admin.export_excel trip_id=%s members=%s", trip_id, len(members))
        
        # Reset file position for download
        excel_file.seek(0)
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    finally:
        try:
            next(db_gen)
        except StopIteration:
            pass


@router.get("/trip/{trip_id}/edit")
async def edit_trip_page(request: Request, trip_id: int):
    """Show edit trip form."""
    # Check if request is from Telegram WebApp
    error_response = require_telegram_webapp(request, Config.BOT_USERNAME)
    if error_response:
        return error_response
    
    db_gen = get_db()
    db: Session = next(db_gen)
    
    try:
        trip = db.query(Trip).filter(Trip.id == trip_id).first()
        if not trip:
            raise HTTPException(status_code=404, detail="Trip not found")
        
        return templates.TemplateResponse(
            "edit_trip.html",
            {
                "request": request,
                "trip": trip,
            }
        )
    finally:
        try:
            next(db_gen)
        except StopIteration:
            pass


@router.post("/api/trip/{trip_id}/update")
async def update_trip(request: Request, trip_id: int):
    """Update trip properties."""
    _require_admin(request)
    
    db_gen = get_db()
    db: Session = next(db_gen)
    
    try:
        trip = db.query(Trip).filter(Trip.id == trip_id).first()
        if not trip:
            raise HTTPException(status_code=404, detail="Trip not found")
        
        # Get form data
        form_data = await request.json()
        
        # Update fields
        if 'name' in form_data:
            trip.name = form_data['name'].strip()
        if 'price' in form_data:
            trip.price = int(form_data['price'])
        if 'participant_limit' in form_data:
            limit_value = form_data['participant_limit']
            trip.participant_limit = int(limit_value) if limit_value and str(limit_value).strip() else None
        if 'card_info' in form_data:
            trip.card_info = form_data['card_info'].strip() or None
        if 'agreement_text' in form_data:
            trip.agreement_text = form_data['agreement_text'].strip() or None
        
        db.commit()
        db.refresh(trip)
        
        logging.info("admin.trip_updated trip_id=%s", trip_id)
        
        return JSONResponse({
            "success": True,
            "message": "Trip updated successfully",
            "trip": {
                "id": trip.id,
                "name": trip.name,
                "price": trip.price,
                "participant_limit": trip.participant_limit,
                "card_info": trip.card_info,
                "agreement_text": trip.agreement_text,
            }
        })
        
    except ValueError as e:
        return JSONResponse(
            status_code=400,
            content={"success": False, "message": f"Invalid input: {str(e)}"}
        )
    except Exception as e:
        db.rollback()
        logging.error(f"Error updating trip {trip_id}: {e}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={"success": False, "message": "Failed to update trip"}
        )
    finally:
        try:
            next(db_gen)
        except StopIteration:
            pass
